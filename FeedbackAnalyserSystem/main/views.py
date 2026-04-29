from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.db.models import Avg, Q, Count
from django.db.models.functions import TruncDate
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from django.db import transaction
from django.contrib import messages
from django.core.exceptions import PermissionDenied

from users.models import Platform, SentimentAnchor
from users.models import Feedback, SentimanetAnalyze, EmployeeProfile, NotificationSetting
from users.forms import EmployeeCreationForm, EmployeeEditForm
from users.models import CompanyMember
from django.http import HttpResponse
from users.forms import SentimentAnchorForm
from users.utils import get_analyzer
from django.views.decorators.http import require_POST
import openpyxl
from django.core.mail import send_mail
from users.decorators import require_role

from .utils import get_active_company
from users.models import Notification


@login_required(login_url='/users/login/')
def dashboard_view(request):
    user_company = get_active_company(request)

    if not user_company:
        raise PermissionDenied("Вы не состоите ни в одной компании.")
        
    today = timezone.now().date()
    fourteen_days_ago = today - timedelta(days=14)
    
    # ==========================================
    # 1. КЛЮЧЕВЫЕ МЕТРИКИ (KPI)
    # ==========================================
    total_reviews_today = Feedback.objects.filter(
        company=user_company, 
        created_at__date=today
    ).count()
    
    avg_sentiment = SentimanetAnalyze.objects.filter(
        feedback__company=user_company, 
        type='FINAL'                    
    ).aggregate(avg_val=Avg('value'))['avg_val']
    
    avg_index = int(avg_sentiment * 100) if avg_sentiment is not None else 0
        
    critical_count = SentimanetAnalyze.objects.filter(
        Q(value__lt=-0.5) | Q(meta_data__status__icontains="SARCASM"),
        feedback__company=user_company,
        type='FINAL'
    ).count()

    # ==========================================
    # 2. ПОДГОТОВКА ДАННЫХ ДЛЯ ГРАФИКОВ
    # ==========================================
    base_qs = SentimanetAnalyze.objects.filter(
        type='FINAL',                   
        feedback__company=user_company, 
        feedback__created_at__gte=fourteen_days_ago
    )

    # --- Кольцевая диаграмма (Распределение) ---
    pos_count = base_qs.filter(value__gt=0.2).count()
    neg_count = base_qs.filter(value__lt=-0.2).count()
    neu_count = base_qs.filter(value__gte=-0.2, value__lte=0.2).count()

    distribution_data = {
        'labels': ['Нейтральные', 'Позитивные', 'Негативные'],
        'values': [neu_count, pos_count, neg_count]
    }

    # --- Линейный график (Тренд тональности) ---
    daily_stats = base_qs.annotate(date=TruncDate('feedback__created_at')) \
        .values('date') \
        .annotate(
            pos_count=Count('id', filter=Q(value__gt=0.2)),
            neg_count=Count('id', filter=Q(value__lt=-0.2))
        ).order_by('date')

    trend_labels = []
    trend_positive = []
    trend_negative = []

    for stat in daily_stats:
        if stat['date']:
            formatted_date = stat['date'].strftime('%d %b')
            trend_labels.append(formatted_date)
            trend_positive.append(stat['pos_count'])
            trend_negative.append(stat['neg_count'])

    trend_data = {
        'labels': trend_labels,
        'positive': trend_positive,
        'negative': trend_negative
    }

    context = {
        'page_title': 'Дашборд',
        'total_reviews': total_reviews_today,
        'avg_index': avg_index,
        'critical_count': critical_count,
        'trend_data': trend_data,
        'distribution_data': distribution_data,
    }

    return render(request, 'main/dashboard.html', context)

@login_required(login_url='/users/login/')
@require_role(['owner', 'admin', 'manager'])
def feedback_view(request):
    user_company = get_active_company(request)

    all_feedbacks = Feedback.objects.filter(company=user_company)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    selected_source = request.GET.get('source_filter')
    selected_tag = request.GET.get('tag_filter')
    selected_search = request.GET.get('search_query')
    selected_rating = request.GET.get('rating_filter')

    if date_from:
        all_feedbacks = all_feedbacks.filter(created_at__date__gte=date_from)
        
    if date_to:
        all_feedbacks = all_feedbacks.filter(created_at__date__lte=date_to)
        
    if selected_source:
        all_feedbacks = all_feedbacks.filter(platform__name=selected_source)
    if selected_tag:
        all_feedbacks = all_feedbacks.filter(category__icontains=selected_tag)
    if selected_search:
        all_feedbacks = all_feedbacks.filter(
            Q(text__icontains=selected_search) | 
            Q(external_id__icontains=selected_search)
        )
    if selected_rating and selected_rating.isdigit():
        all_feedbacks = all_feedbacks.filter(rating__gte=float(selected_rating))

    all_feedbacks = all_feedbacks.order_by('-created_at')

    paginator = Paginator(all_feedbacks, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    feedback_ids = [feedback.id for feedback in page_obj]

    sentiments = SentimanetAnalyze.objects.filter(
        feedback_id__in=feedback_ids, 
        type='FINAL'
    ).values('feedback_id', 'value')

    sentiment_map = {item['feedback_id']: item['value'] for item in sentiments}

    for feedback in page_obj:
        feedback.sentiment_value = sentiment_map.get(feedback.id, 0.0)

    context = {
        'page_obj': page_obj,
        'platforms': Platform.objects.filter(is_active=True)
    }

    return render(request, 'main/feed.html', context)

@login_required(login_url='/users/login/')
@require_role(['owner', 'admin'])
def config(request):
    if not hasattr(request.user, 'profile'):
        return render(request, 'main/config.html', {'page_obj': [], 'form': EmployeeCreationForm()})

    user_company = get_active_company(request)
    if not user_company:
        raise PermissionDenied("Вы не состоите ни в одной компании.")

    available_roles = [
        ('owner', 'Владелец'),
        ('admin', 'Администратор'),
        ('manager', 'Менеджер'),
    ]

    # Обработка POST-запросов (маршрутизация по параметру action)
    if request.method == 'POST':
        action = request.POST.get('action', 'create_user')

        # 1. Создание нового пользователя
        if action == 'create_user':
            form = EmployeeCreationForm(request.POST)
            if form.is_valid():
                with transaction.atomic():
                    new_user = form.save()
                    EmployeeProfile.objects.create(
                        user=new_user,
                        phone=form.cleaned_data.get('phone', ''),
                        slug=new_user.username 
                    )
                    CompanyMember.objects.create(
                        user=new_user,
                        company=user_company,
                        role=form.cleaned_data.get('role', 'manager')
                    )
                messages.success(request, "Новый сотрудник успешно создан.")
                return redirect('config')
        
        # 2. Обновление роли существующего сотрудника
        elif action == 'update_role':
            current_member = CompanyMember.objects.get(user=request.user, company=user_company)
            if current_member.role != 'owner':
                messages.error(request, "Доступ запрещен. Только владелец может изменять роли.")
                return redirect('config')

            user_id = request.POST.get('user_id') 
            new_role = request.POST.get('new_role')

            try:
                member_to_update = CompanyMember.objects.get(user_id=user_id, company=user_company)
                
                if member_to_update.user == request.user and new_role != 'owner':
                    owners_count = CompanyMember.objects.filter(company=user_company, role='owner').count()
                    if owners_count <= 1:
                        messages.error(request, "Ошибка: Невозможно понизить единственного владельца.")
                        return redirect('config')

                member_to_update.role = new_role
                member_to_update.save()
                messages.success(request, f"Роль {member_to_update.user.username} обновлена.")
                
            except CompanyMember.DoesNotExist:
                messages.error(request, "Сотрудник не найден.")
                
            return redirect('config')

        # 3. Добавление существующего системного пользователя в компанию
        elif action == 'add_user':
            new_user_id = request.POST.get('new_user_id')
            new_role = request.POST.get('new_role', 'manager')

            try:
                user_to_add = User.objects.get(id=new_user_id)
                if CompanyMember.objects.filter(user=user_to_add, company=user_company).exists():
                    messages.warning(request, "Данный пользователь уже состоит в компании.")
                else:
                    CompanyMember.objects.create(user=user_to_add, company=user_company, role=new_role)
                    messages.success(request, f"Пользователь {user_to_add.username} добавлен в компанию.")
            except User.DoesNotExist:
                messages.error(request, "Ошибка: Пользователь не найден в системе.")
            return redirect('config')

    # Инициализация формы, если POST для 'create_user' был невалиден или это GET-запрос
    if request.method == 'GET' or (request.method == 'POST' and request.POST.get('action') != 'create_user'):
        form = EmployeeCreationForm()

    # --- Обработка GET-запроса (Фильтрация, сортировка, поиск) ---
    all_users = User.objects.filter(companies=user_company)

    # Фильтры для текущих сотрудников
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    sort_date = request.GET.get('sort_date', '-date_joined')

    if search_query:
        all_users = all_users.filter(
            Q(username__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(profile__phone__icontains=search_query)
        )
    if role_filter in ['owner', 'admin', 'manager']:
        all_users = all_users.filter(
            memberships__company=user_company, 
            memberships__role=role_filter
        )

    if status_filter == 'active':
        all_users = all_users.filter(is_active=True)
    elif status_filter == 'inactive':
        all_users = all_users.filter(is_active=False)

    if sort_date in ['date_joined', '-date_joined']:
        all_users = all_users.order_by(sort_date)
    else:
        all_users = all_users.order_by('-date_joined')

    paginator = Paginator(all_users, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Логика поиска системных пользователей (не состоящих в компании)
    search_sys_query = request.GET.get('search_user', '').strip()
    found_users = []
    
    if search_sys_query:
        existing_user_ids = all_users.values_list('id', flat=True)
        found_users = User.objects.filter(
            Q(username__icontains=search_sys_query) | Q(email__icontains=search_sys_query)
        ).exclude(id__in=existing_user_ids)[:5]

    context = {
        'page_obj': page_obj,
        'form': form,
        'available_roles': available_roles,
        'search_sys_query': search_sys_query,
        'found_users': found_users,
    }
    return render(request, 'main/config.html', context)

@login_required(login_url='/users/login/')
@require_role(['owner', 'admin'])
def delete_employee(request, user_id):
    if request.method == 'POST':
        user_company = get_active_company(request)

        if user_id == request.user.id:
            return redirect('config')

        membership = get_object_or_404(CompanyMember, user_id=user_id, company=user_company)
        membership.delete()

    return redirect('config')


@login_required(login_url='/users/login/')
@require_role(['owner', 'admin'])
def edit_employee(request, user_id):
    user_company = get_active_company(request)

    if not user_company:
        raise PermissionDenied("Вы не состоите ни в одной компании.")
    
    user_to_edit = get_object_or_404(User, id=user_id, companies=user_company)

    if request.method == 'POST':
        form = EmployeeEditForm(request.POST, instance=user_to_edit)
        if form.is_valid():
            with transaction.atomic():
                form.save()
            return redirect('config')
    else:
        form = EmployeeEditForm(instance=user_to_edit)

    context = {
        'form': form,
        'user_to_edit': user_to_edit
    }
    return render(request, 'main/edit_user.html', context)

@login_required(login_url='/users/login/')
def dictionary_view(request):
    if request.method == 'POST':
        form = SentimentAnchorForm(request.POST)
        if form.is_valid():
            form.save()
            
            try:
                analyzer = get_analyzer()
                analyzer.reload_anchors()
            except NameError:
                print("Функция get_analyzer не импортирована. Перезапустите сервер вручную.")
                
            return redirect('dictionary')
    else:
        form = SentimentAnchorForm()

    anchors = SentimentAnchor.objects.all().order_by('-created_at')

    search_query = request.GET.get('search', '')
    sentiment_filter = request.GET.get('sentiment', '')
    language_filter = request.GET.get('language', '')  
    status_filter = request.GET.get('status', '')

    if search_query:
        anchors = anchors.filter(text__icontains=search_query)
    
    if sentiment_filter:
        anchors = anchors.filter(sentiment=sentiment_filter)

    if language_filter:
        anchors = anchors.filter(language=language_filter)
        
    if status_filter == 'active':
        anchors = anchors.filter(is_active=True)
    elif status_filter == 'inactive':
        anchors = anchors.filter(is_active=False)

    paginator = Paginator(anchors, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sentiment_filter': sentiment_filter,
        'language_filter': language_filter, 
        'status_filter': status_filter,
        'form': form, 
    }
    return render(request, 'main/dictionary.html', context)

@login_required(login_url='/users/login/')
@require_POST
def delete_anchor(request, pk):
    anchor = get_object_or_404(SentimentAnchor, pk=pk)
    anchor.delete()
    
    try:
        analyzer = get_analyzer()
        analyzer.reload_anchors()
    except Exception as e:
        print(f"Ошибка обновления словаря: {e}")
        
    return redirect('dictionary')

@login_required(login_url='/users/login/')
@require_POST
def edit_anchor(request, pk):
    anchor = get_object_or_404(SentimentAnchor, pk=pk)
    
    form = SentimentAnchorForm(request.POST, instance=anchor)
    
    if form.is_valid():
        form.save()
        
        try:
            analyzer = get_analyzer()
            analyzer.reload_anchors()
        except Exception as e:
            print(f"Ошибка обновления словаря: {e}")
            
    return redirect('dictionary')



@login_required(login_url='/users/login/')
@require_POST
def import_anchors(request):
    excel_file = request.FILES.get('excel_file')
    
    if not excel_file or not excel_file.name.endswith('.xlsx'):
        messages.error(request, "Ошибка: Пожалуйста, загрузите файл формата .xlsx")
        return redirect('dictionary')

    try:
        # Читаем Excel файл напрямую из оперативной памяти (без сохранения на диск)
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        
        new_anchors = []
        valid_sentiments = ['POSITIVE', 'NEGATIVE']
        
        # Начинаем со 2-й строки (min_row=2), так как в 1-й строке будут заголовки
        for row in sheet.iter_rows(min_row=2, values_only=True):
            text = row[0]
            sentiment = row[1]
            language = row[2] if len(row) > 2 and row[2] else 'ru' 
            
            # Если строка не пустая и тональность указана верно
            if text and sentiment in valid_sentiments:
                new_anchors.append(SentimentAnchor(
                    text=str(text).strip(),
                    sentiment=str(sentiment).strip().upper(),
                    language=str(language).strip().lower()
                ))

        # Массовое сохранение (ignore_conflicts=True пропустит слова, которые уже есть в базе)
        if new_anchors:
            SentimentAnchor.objects.bulk_create(new_anchors, ignore_conflicts=True)
            
            analyzer = get_analyzer()
            analyzer.reload_anchors()
            
            messages.success(request, f"Успешно обработано {len(new_anchors)} слов. Словарь ИИ обновлен.")
        else:
            messages.warning(request, "Файл пуст или данные не соответствуют формату.")
            
    except Exception as e:
        messages.error(request, f"Произошла ошибка при чтении файла: {e}")

    return redirect('dictionary')

@login_required(login_url='/users/login/')
def profile_view(request):
    user = request.user
    active_comp = get_active_company(request)
    
    has_any_company = CompanyMember.objects.filter(user=user).exists()
    
    if active_comp:
        company_label = active_comp.name
    else:
        company_label = "Компания не выбрана"
    
    if hasattr(user, 'profile'):
        phone = user.profile.phone if user.profile.phone else "Не указан"
    else:
        phone = "Не указан"

    context = {
        'company': company_label,
        'phone': phone,
        'is_workspace_selection': active_comp is None or not has_any_company,
    }
    
    return render(request, 'main/profile.html', context)

@login_required(login_url='/users/login/')
@require_role(['owner', 'admin', 'manager']) 
@require_POST
def override_sentiment(request, feedback_id):
    active_company = get_active_company(request)

    feedback = get_object_or_404(Feedback, id=feedback_id, company=active_company)

    new_value = request.POST.get('new_value') #ожидаем от -1.0 до 1.0

    if new_value is None:
        return JsonResponse({'error': 'Не передано значение new_value'}, status=400)
        
    try:
        new_value = float(new_value)
    except ValueError:
        return JsonResponse({'error': 'Значение должно быть числом'}, status=400)
    
    final_sentiment, created = SentimanetAnalyze.objects.get_or_create(
        feedback=feedback,
        type='FINAL',
        defaults={'value': new_value}
    )

    # Обновляем метрики в зависимости от выбора
    final_sentiment.value = new_value
    final_sentiment.positive_val = 1.0 if new_value > 0 else 0.0
    final_sentiment.negative_val = -1.0 if new_value < 0 else 0.0 # храним отрицательное
    final_sentiment.neutral_val = 1.0 if new_value == 0 else 0.0
    
    final_sentiment.is_manual = True
    final_sentiment.edited_by = request.user
    final_sentiment.save()
    
    return JsonResponse({
        'status': 'success', 
        'message': 'Оценка успешно изменена',
        'new_value': new_value
    })


@login_required(login_url='/users/login/')
@require_role(['owner', 'admin']) 
def notifications_settings_view(request):
    active_company = get_active_company(request)

    settings_obj, created = NotificationSetting.objects.get_or_create(company=active_company)

    if request.method == 'POST':
        settings_obj.is_in_app_enabled = request.POST.get('is_in_app_enabled') == 'on'
        settings_obj.is_telegram_enabled = request.POST.get('is_telegram_enabled') == 'on'
        settings_obj.is_email_enabled = request.POST.get('is_email_enabled') == 'on'
        
        threshold = request.POST.get('critical_threshold')
        if threshold:
            try:
                settings_obj.critical_threshold = float(threshold)
            except ValueError:
                messages.error(request, "Неверный формат числа для порога.")
                return redirect('notifications')
        
        settings_obj.telegram_bot_token = request.POST.get('telegram_bot_token', '')
        settings_obj.telegram_chat_id = request.POST.get('telegram_chat_id', '')
        settings_obj.custom_emails = request.POST.get('custom_emails', '')
        settings_obj.alert_template = request.POST.get('alert_template', '')

        settings_obj.save()
        
        messages.success(request, "Конфигурация уведомлений успешно сохранена!")
        return redirect('notifications')

    return render(request, 'users/notifications_settings.html', {'settings': settings_obj})

@login_required(login_url='/users/login/')
def export_feedback_excel(request):
    active_company = get_active_company(request)
    feedbacks = Feedback.objects.filter(company=active_company).select_related('platform')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отзывы"

    headers = ['Платформа', 'Оценка', 'Текст отзыва', 'Внешний ID']
    ws.append(headers)

    for fb in feedbacks:
        platform_name = fb.platform.name if fb.platform else "Неизвестно"
        
        ws.append([
            platform_name,
            fb.rating,
            fb.text,
            fb.external_id
        ])

    ws.column_dimensions['D'].width = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Feedbacks_Export.xlsx"'
    
    wb.save(response)

    return response


@login_required(login_url='/users/login/')
def select_company_view(request):
    user_memberships = CompanyMember.objects.filter(user=request.user).select_related('company')

    if not user_memberships.exists():
        return render(request, 'users/no_companies.html', {'is_workspace_selection': True})

    if user_memberships.count() == 1:
        company_id = user_memberships.first().company.id
        request.session['active_company_id'] = company_id
        return redirect('dashboard')

    return render(request, 'users/select_company.html', {
        'memberships': user_memberships,
        'is_workspace_selection': True 
    })

@login_required(login_url='/users/login/')
def set_active_company(request, company_id):
    has_access = CompanyMember.objects.filter(user=request.user, company_id=company_id).exists()
    
    if has_access:
        # ЗАПИСЫВАЕМ В СЕССИЮ! Это самое главное.
        request.session['active_company_id'] = company_id
        return redirect('dashboard')
    else:
        messages.error(request, "У вас нет доступа к этой системе.")
        return redirect('select_company')
    
@login_required(login_url='/users/login/')
@require_role(['owner']) 
def manage_permissions_view(request):
    active_company = get_active_company(request)
    
    available_roles = [
        ('owner', 'Владелец'),
        ('admin', 'Администратор'),
        ('manager', 'Менеджер'),
    ]

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_role':
            member_id = request.POST.get('member_id')
            new_role = request.POST.get('new_role')

            try:
                member_to_update = CompanyMember.objects.get(id=member_id, company=active_company)
                
                if member_to_update.user == request.user and new_role != 'owner':
                    owners_count = CompanyMember.objects.filter(company=active_company, role='owner').count()
                    if owners_count <= 1:
                        messages.error(request, "Ошибка: Невозможно понизить единственного владельца.")
                        return redirect('manage_permissions')

                member_to_update.role = new_role
                member_to_update.save()
                messages.success(request, f"Роль {member_to_update.user.username} обновлена.")
                
            except CompanyMember.DoesNotExist:
                messages.error(request, "Сотрудник не найден.")

        elif action == 'add_user':
            new_user_id = request.POST.get('new_user_id')
            new_role = request.POST.get('new_role', 'manager')

            try:
                user_to_add = User.objects.get(id=new_user_id)
                if CompanyMember.objects.filter(user=user_to_add, company=active_company).exists():
                    messages.warning(request, "Данный пользователь уже состоит в компании.")
                else:
                    CompanyMember.objects.create(user=user_to_add, company=active_company, role=new_role)
                    messages.success(request, f"Пользователь {user_to_add.username} добавлен в компанию.")
            except User.DoesNotExist:
                messages.error(request, "Ошибка: Пользователь не найден в системе.")

        return redirect('manage_permissions')

    members = CompanyMember.objects.filter(company=active_company).select_related('user')
    
    search_query = request.GET.get('search_user', '').strip()
    found_users = []
    
    if search_query:
        existing_user_ids = members.values_list('user_id', flat=True)
        found_users = User.objects.filter(
            Q(username__icontains=search_query) | Q(email__icontains=search_query)
        ).exclude(id__in=existing_user_ids)[:5]

    context = {
        'members': members,
        'available_roles': available_roles,
        'search_query': search_query,
        'found_users': found_users,
    }
    return render(request, 'users/manage_permissions.html', context)


@login_required(login_url='/users/login/')
@require_role(['owner', 'admin', 'manager'])
def reports_view(request):
    active_company = get_active_company(request)

    # 1. Базовый QuerySet для компании
    feedbacks = Feedback.objects.filter(company=active_company)
    analyses = SentimanetAnalyze.objects.filter(feedback__company=active_company)

    # 2. Общие показатели
    total_feedbacks = feedbacks.count()
    avg_rating = feedbacks.aggregate(Avg('rating'))['rating__avg'] or 0.0
    avg_sentiment = analyses.aggregate(Avg('value'))['value__avg'] or 0.0

    # 3. Распределение по тональности (используем пороги: >0.2 позитив, <-0.2 негатив)
    final_analyses = analyses.filter(type='FINAL')

    positive_count = final_analyses.filter(value__gt=0.2).count()
    negative_count = final_analyses.filter(value__lt=-0.2).count()
    neutral_count = final_analyses.filter(value__range=(-0.2, 0.2)).count()

    # 4. Статистика по источникам (Платформы)
    platform_stats = feedbacks.values('platform__name') \
        .annotate(count=Count('id')) \
        .order_by('-count')
    
    category_stats = feedbacks.values('category') \
        .annotate(
            count=Count('id', distinct=True),
            avg_sentiment=Avg('analyses__value')
        ) \
        .order_by('-count')

    # 5. Временной ряд (последние 30 дней)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    daily_stats = feedbacks.filter(created_at__gte=thirty_days_ago) \
        .annotate(date=TruncDate('created_at')) \
        .values('date') \
        .annotate(count=Count('id')) \
        .order_by('date')

    context = {
        'total_feedbacks': total_feedbacks,
        'avg_rating': round(avg_rating, 2),
        'avg_sentiment': round(avg_sentiment, 2),
        'positive_count': positive_count,
        'negative_count': negative_count,
        'neutral_count': neutral_count,
        'platform_stats': platform_stats,
        'category_stats': category_stats,
        'daily_stats': list(daily_stats),
    }
    
    return render(request, 'main/reports.html', context)

@login_required(login_url='/users/login/')
def mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    
    return redirect(request.META.get('HTTP_REFERER', '/'))